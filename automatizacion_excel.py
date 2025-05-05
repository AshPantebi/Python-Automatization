"""Automatizador de reportes de excel adaptable para cualquier conjunto de datos"""
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import matplotlib as mpl

# 1- Importar el archivo con los datos
df = pd.read_csv("supermarket_sales new.csv")

# Creando nuevas variables (Opcional)
df['Subtotal'] = (df['Unit price'] * df['Quantity']).round(2)
df['Total'] = (df['Subtotal'] + df['Tax 5%']).round(2)

# 2- Procesando los datos para reportar
# 2.1 Definiendo el contenido del resumen estadístico (Adaptable a gusto del cliente)
resumen_estadistico = {
    'Total de Ventas'      : df['Total'].sum(),
    'Productos Vendidos'   : df['Quantity'].sum(),
    'Número de Clientes'   : len(df['Invoice ID'].unique()),
    'Promedio de Ventas'   : df['Total'].mean(),
    'Mediana de Ventas'    : df['Total'].median(),
    'Dispersión de Ventas' : df['Total'].std(),
    'Mayor Venta'          : df['Total'].max(),
    'Menor Venta'          : df['Total'].min(),
    'Total Impuestos (5%)' : df['Tax 5%'].sum()
}

# Tabulando resumen estadístico
resumen_estadistico = pd.Series(resumen_estadistico, name='Valor')
resumen_estadistico = pd.DataFrame(resumen_estadistico).round(2)
resumen_estadistico.index.name = 'Métrica'

# Aquí abajo puede ir cualquier tabulación de las variables.
# Por ejemplo: Ventas por Línea de Productos vs Género

#tbl1 = df.pivot_table(index=['Product line'], columns=['Gender'], values=['Total'], aggfunc='sum')

# 3- Escribiendo el archivo excel
# Aquí se pueden escribir las hojas que sean para excel
# Lo ideal sería hacer un bucle for o una función
with pd.ExcelWriter('ejemplo.xlsx') as writer:
    df.to_excel(writer, sheet_name='Dataset', index=False) # -> La tabla de datos irá en la primera hoja "Dataset"
    resumen_estadistico.to_excel(writer, sheet_name='Reporte') # -> Reporte en la 2da hoja "Reporte"

# 4- Editando el archivo excel
# 4.1 Cargando el excel y las hojas a editar
wb = openpyxl.load_workbook('ejemplo.xlsx')
ws1 = wb['Dataset']
ws2 = wb['Reporte']

# 4.2 Personalizando hoja 'Dataset'
for columna in range(1, ws1.max_column + 1):
    # Configurando ancho de todas las celdas en cada columna
    ws1.column_dimensions[get_column_letter(columna)].width = 20

    # Esto si personaliza el color y las letras de las celdas
    celda = ws1.cell(row=1, column=columna)
    celda.fill = PatternFill(start_color='006A71',fill_type='solid') # -> Aquí se puede cambiar el color de las celdas
    celda.font = Font(bold=True, color='EFEFEF', size=12) # -> Aquí se puede cambiar el color de las letras

    # Personalizando el resto de la tabla
    for fila in range(1, ws1.max_row + 1):
        celda = ws1.cell(row=fila, column=columna)
        celda.alignment = Alignment(horizontal='center') # -> Aquí se puede cambiar la orientación del texto

# 4.3 Personalizando el reporte
# 4.3.1 Personalizando la tabla resumen
for columna in range(1, ws2.max_column + 1):
    # Configurando ancho de todas las celdas en cada columna
    ws2.column_dimensions[get_column_letter(columna)].width = 20

    # Esto si personaliza el color y las letras de las celdas
    celda = ws2.cell(row=1, column=columna)
    celda.fill = PatternFill(start_color='006A71',fill_type='solid') # -> Aquí se puede cambiar el color de las celdas
    celda.font = Font(bold=True, color='EFEFEF', size=12) # -> Aquí se puede cambiar el color de las letras
    celda.alignment = Alignment(horizontal='center')

    # Personalizando el resto de celdas
    for fila in range(2, ws1.max_row + 1):
        celda = ws2.cell(row=fila, column=columna)
        celda.alignment = Alignment(horizontal='left') # -> Aquí se puede cambiar la orientación del texto

# 4.3.2 Creando los gráficos para el reporte
# Creando la imagen y la distribución de los gráficos
fig = plt.figure(layout='constrained', figsize=(10,6)) # -> Dimensiones de la imagen
gs = plt.GridSpec(2, 3, figure=fig) # -> La cantidad de gráficos base (2 filas y 3 columnas de gráficos)

# Definiendo los parámetros para graficar
variables = ['Gender', 'Branch', 'City', 'Customer type'] # -> Variables a graficar
titulos   = ['Género', 'Supermercado', 'Ciudad', 'Tipo de Clientes'] # -> Titulo para cada gráfico
colores   = list(mpl.colors.TABLEAU_COLORS) # -> Lista de colores para usar en cada gráfico

# Creando los primeros 4 gráficos
for grafico, variable, titulo in zip(range(4), variables, titulos):
    # Posición de cada gráfico
    ax = fig.add_subplot(gs[grafico])

    # Configuración de cada uno de los primeros 4 graficos
    configuracion = {
        'kind'  : 'bar', 
        'xlabel': '', 
        'ylabel': 'Total', 
        'rot'   : 0, 
        'title' : f'Ventas por {variable}', 
        'color' : colores, 
        'ax'    : ax
    }

    # Generando los gráficos
    df.groupby(variable)['Total'].agg('sum').plot(**configuracion)

# Creando el último gráfico que ocupará dos espacios dentro de la grilla 2x3
ax = fig.add_subplot(gs[4:]) # -> Posición

# Configuración
configuracion = {
    'kind'         : 'pie',
    'ax'           : ax,
    'autopct'      : '%1.1f%%',
    'legend'       : True,
    'title'        : 'Ventas por Línea de Productos',
    'ylabel'       : '',
    'labeldistance': None
}

# Creación
df.groupby('Product line')['Total'].agg('sum').plot(**configuracion)

# Leyenda
ax.legend(bbox_to_anchor=(1, 1.02), loc='upper left')

# Guardando la imagen en la memoria ram para evitar guardar en el disco
buffer = io.BytesIO()
fig.savefig(buffer, format='png')
buffer.seek(0)

# Colocando la imagen con los gráficos en el reporte
img = Image(buffer)
ws2.add_image(img, 'D1')

# Guardando el archivo y eliminando la imagen de la memoria
wb.save('ejemplo.xlsx')
plt.close()
buffer.close()

print('Se ha generado el reporte')
# Fin del archivo (EOF)